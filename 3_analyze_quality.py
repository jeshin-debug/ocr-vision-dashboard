# -*- coding: utf-8 -*-
"""
Project 2: OCR Dashboard Practice
Step 3: OCR Quality Analysis and Excel Reporting Script
Filename: 3_analyze_quality.py
"""

import os
import sys
import json
import pandas as pd
import numpy as np

# Set console output encoding to UTF-8 for Korean support immediately
if sys.platform.startswith('win'):
    import subprocess
    try:
        subprocess.run('chcp 65001', shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except:
        pass
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass

# Ensure openpyxl is installed for writing styled Excel sheets
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[INFO] Excel Setup: Installing 'openpyxl' library...")
    import subprocess
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        print("[INFO] 'openpyxl' installation complete!")
    except Exception as e:
        print(f"[WARNING] 'openpyxl' installation failed: {e}. Basic Excel saving will be used.")

# Paths
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
GT_CSV_PATH = os.path.join(BASE_DIR, "data", "source_structured", "ground_truth_multimodal_240.csv")
OCR_CSV_PATH = os.path.join(BASE_DIR, "data", "ocr", "ocr_extracted_raw.csv")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
SUMMARY_TXT_PATH = os.path.join(REPORTS_DIR, "ocr_quality_summary.txt")
REPORT_XLSX_PATH = os.path.join(REPORTS_DIR, "ocr_quality_report.xlsx")

def safe_parse_json(val):
    if pd.isna(val) or not val:
        return {}
    try:
        return json.loads(val)
    except:
        return {}

def main():
    print("=" * 60)
    print(" Project 2: OCR 추출 품질 및 정확도 전수 평가")
    print("=" * 60)
    print(f"OCR 파일: {OCR_CSV_PATH}")
    print(f"정답 파일: {GT_CSV_PATH}")
    
    if not os.path.exists(GT_CSV_PATH):
        print(f"[오류] 정답 CSV 파일이 존재하지 않습니다: {GT_CSV_PATH}")
        sys.exit(1)
    if not os.path.exists(OCR_CSV_PATH):
        print(f"[오류] OCR 결과 CSV 파일이 존재하지 않습니다: {OCR_CSV_PATH}\n(먼저 2_ocr_extractor.py를 실행해주세요.)")
        sys.exit(1)
        
    # Read files
    df_gt = pd.read_csv(GT_CSV_PATH, encoding='utf-8')
    df_ocr = pd.read_csv(OCR_CSV_PATH, encoding='utf-8')
    
    # Merge datasets
    df_merged = pd.merge(df_ocr, df_gt, on="record_id", suffixes=('_ocr', '_gt'))
    total_records = len(df_merged)
    
    print(f"-> 총 {total_records}개의 레코드가 매칭되어 분석을 시작합니다.")
    
    # Initialize metric variables
    # 1. Document Type Classification
    doc_type_matches = (df_merged['document_type_ocr'] == df_merged['document_type_gt']).sum()
    doc_type_accuracy = doc_type_matches / total_records if total_records > 0 else 0
    
    # Split into receipts and surveys
    df_receipts = df_merged[df_merged['document_type_gt'] == 'receipt'].copy()
    df_surveys = df_merged[df_merged['document_type_gt'] == 'survey'].copy()
    
    # 2. Receipt Amount Accuracy
    total_receipts = len(df_receipts)
    receipt_amount_matches = 0
    if total_receipts > 0:
        # Match only when both are numeric and equal
        receipt_amount_matches = ((df_receipts['extracted_amount'] == df_receipts['total_amount'])).sum()
        receipt_amount_accuracy = receipt_amount_matches / total_receipts
    else:
        receipt_amount_accuracy = 0
        
    # 3. Survey Scores Accuracy
    total_surveys = len(df_surveys)
    survey_scores_evaluated = 0
    survey_scores_matches = 0
    
    survey_details_rows = []
    
    for idx, row in df_surveys.iterrows():
        ocr_scores = safe_parse_json(row['extracted_scores'])
        
        sat_ocr = ocr_scores.get('satisfaction_score')
        usa_ocr = ocr_scores.get('usability_score')
        spd_ocr = ocr_scores.get('speed_score')
        
        sat_gt = row['satisfaction_score']
        usa_gt = row['usability_score']
        spd_gt = row['speed_score']
        
        sat_match = (sat_ocr == sat_gt) if sat_ocr is not None else False
        usa_match = (usa_ocr == usa_gt) if usa_ocr is not None else False
        spd_match = (spd_ocr == spd_gt) if spd_ocr is not None else False
        
        survey_scores_evaluated += 3
        if sat_match: survey_scores_matches += 1
        if usa_match: survey_scores_matches += 1
        if spd_match: survey_scores_matches += 1
        
        survey_details_rows.append({
            "record_id": row['record_id'],
            "department": row['respondent_dept'],
            "satisfaction_gt": sat_gt,
            "satisfaction_ocr": sat_ocr,
            "satisfaction_match": "일치" if sat_match else "불일치",
            "usability_gt": usa_gt,
            "usability_ocr": usa_ocr,
            "usability_match": "일치" if usa_match else "불일치",
            "speed_gt": spd_gt,
            "speed_ocr": spd_ocr,
            "speed_match": "일치" if spd_match else "불일치",
        })
        
    survey_score_accuracy = survey_scores_matches / survey_scores_evaluated if survey_scores_evaluated > 0 else 0
    
    # 4. Survey Note Extraction success rate
    survey_note_non_null = 0
    survey_note_matches = 0
    if total_surveys > 0:
        # Note is extracted successfully if not NaN and not empty
        survey_note_non_null = df_surveys['extracted_note'].notna().sum()
        # Matches if equal
        survey_note_matches = (df_surveys['extracted_note'] == df_surveys['handwritten_note']).sum()
        survey_note_extraction_rate = survey_note_non_null / total_surveys
        survey_note_match_rate = survey_note_matches / total_surveys
    else:
        survey_note_extraction_rate = 0
        survey_note_match_rate = 0
        
    # Analyze by quality categories (Resolution and Noise)
    # Low resolution accuracy
    df_low_res = df_merged[df_merged['is_low_resolution'] == True]
    df_normal_res = df_merged[df_merged['is_low_resolution'] == False]
    
    low_res_avg_conf = df_low_res['confidence'].mean() if len(df_low_res) > 0 else 0
    normal_res_avg_conf = df_normal_res['confidence'].mean() if len(df_normal_res) > 0 else 0
    
    # Print and Save txt Summary
    summary_lines = [
        "======================================================================",
        "  [품질 요약 리포트] Project 2 Multimodal OCR 정확도 분석 보고서",
        "======================================================================",
        f"평가 일시: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"총 평가 대상 문서 수: {total_records}건 (영수증: {total_receipts}건, 설문지: {total_surveys}건)",
        "----------------------------------------------------------------------",
        " 항목별 OCR 추출 정확도 지표 (Accuracy & Extraction KPIs)",
        "----------------------------------------------------------------------",
        f"1. 문서 유형 분류 정확도 (Doc Type Classification): {doc_type_accuracy*100:.2f}% ({doc_type_matches}/{total_records}건)",
        f"2. 영수증 금액 추출 정확도 (Receipt Amount Accuracy): {receipt_amount_accuracy*100:.2f}% ({receipt_amount_matches}/{total_receipts}건)",
        f"3. 설문 만족도 점수 정확도 (Survey Score Accuracy): {survey_score_accuracy*100:.2f}% ({survey_scores_matches}/{survey_scores_evaluated}개 항목)",
        f"4. 설문 수기 메모 추출 성공률 (Survey Note Extraction Rate): {survey_note_extraction_rate*100:.2f}% ({survey_note_non_null}/{total_surveys}건)",
        f"   - 수기 메모 100% 매칭 정확도 (Exact Match Rate): {survey_note_match_rate*100:.2f}% ({survey_note_matches}/{total_surveys}건)",
        "----------------------------------------------------------------------",
        " 이미지 품질 상태별 OCR 신뢰도 (Image Quality vs OCR Confidence)",
        "----------------------------------------------------------------------",
        f"- 저해상도 이미지 평균 신뢰도 (Low-Resolution Avg Confidence): {low_res_avg_conf*100:.2f}% ({len(df_low_res)}장)",
        f"- 일반/고해상도 이미지 평균 신뢰도 (Normal-Resolution Avg Confidence): {normal_res_avg_conf*100:.2f}% ({len(df_normal_res)}장)",
        "======================================================================"
    ]
    
    summary_txt = "\n".join(summary_lines)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    with open(SUMMARY_TXT_PATH, "w", encoding="utf-8") as f:
        f.write(summary_txt)
        
    print(summary_txt)
    
    # Create Beautiful Excel Report using openpyxl
    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Styles
        font_title = Font(name="Malgun Gothic", size=16, bold=True, color="1B365D")
        font_section = Font(name="Malgun Gothic", size=12, bold=True, color="1B365D")
        font_header = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Malgun Gothic", size=10)
        font_kpi_num = Font(name="Malgun Gothic", size=20, bold=True, color="107C41")
        font_kpi_label = Font(name="Malgun Gothic", size=9, bold=True, color="555555")
        
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_zebra = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        fill_kpi = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
        fill_match = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        fill_mismatch = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        border_thin = Side(border_style="thin", color="D3D3D3")
        border_double = Side(border_style="double", color="1B365D")
        border_box = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        border_total = Border(top=border_thin, bottom=border_double)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # Sheet 1: 요약 보고서 (Summary Dashboard)
        ws_summary = wb.create_sheet(title="품질 요약 대시보드")
        ws_summary.views.sheetView[0].showGridLines = True
        
        # Title
        ws_summary.merge_cells("A1:G1")
        ws_summary["A1"] = "Project 2: 멀티모달 OCR 정확도 및 품질 평가 보고서"
        ws_summary["A1"].font = font_title
        ws_summary["A1"].alignment = align_left
        ws_summary.row_dimensions[1].height = 40
        
        # Section 1: KPI 카드 영역 (C3:F4)
        kpis = [
            ("전체 분류 정확도", f"{doc_type_accuracy*100:.1f}%", "C3", "C4"),
            ("금액 추출 정확도", f"{receipt_amount_accuracy*100:.1f}%", "D3", "D4"),
            ("설문 점수 정확도", f"{survey_score_accuracy*100:.1f}%", "E3", "E4"),
            ("수기 메모 성공률", f"{survey_note_extraction_rate*100:.1f}%", "F3", "F4")
        ]
        for label, val, c_lbl, c_val in kpis:
            ws_summary[c_lbl] = label
            ws_summary[c_lbl].font = font_kpi_label
            ws_summary[c_lbl].fill = fill_kpi
            ws_summary[c_lbl].alignment = align_center
            ws_summary[c_lbl].border = border_box
            
            ws_summary[c_val] = val
            ws_summary[c_val].font = font_kpi_num
            ws_summary[c_val].fill = fill_kpi
            ws_summary[c_val].alignment = align_center
            ws_summary[c_val].border = border_box
            
        ws_summary.row_dimensions[3].height = 20
        ws_summary.row_dimensions[4].height = 35
        
        # Section 2: 품질 상세 정보 표 (A6:G14)
        ws_summary["A6"] = "📊 상세 항목 평가 지표"
        ws_summary["A6"].font = font_section
        ws_summary.merge_cells("A6:G6")
        
        headers = ["평가 항목", "문서 유형", "전체 대상 건수", "정상 추출 건수", "추출 성공률/정확도", "품질 상태", "평가 기준"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_summary.cell(row=7, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_box
        ws_summary.row_dimensions[7].height = 25
        
        summary_table_data = [
            ["문서 유형 분류", "전체", total_records, doc_type_matches, doc_type_accuracy, "최우수" if doc_type_accuracy > 0.9 else "우수", "영수증/설문지 구분 정합성"],
            ["영수증 금액 추출", "영수증", total_receipts, receipt_amount_matches, receipt_amount_accuracy, "최우수" if receipt_amount_accuracy > 0.9 else "점검필요" if receipt_amount_accuracy < 0.8 else "우수", "영수증 총액(total_amount) 매칭"],
            ["설문 점수 추출", "설문지", survey_scores_evaluated, survey_scores_matches, survey_score_accuracy, "최우수" if survey_score_accuracy > 0.9 else "우수", "만족도/사용성/속도 3개 지표"],
            ["수기 메모 추출", "설문지", total_surveys, survey_note_non_null, survey_note_extraction_rate, "우수" if survey_note_extraction_rate > 0.8 else "점검필요", "수기 메모 텍스트 누락률 검증"],
            ["수기 메모 매칭", "설문지", total_surveys, survey_note_matches, survey_note_match_rate, "우수" if survey_note_match_rate > 0.7 else "점검필요", "수기 메모 내용 100% 일치율"]
        ]
        
        for r_idx, row_data in enumerate(summary_table_data, start=8):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_data
                cell.border = border_box
                
                # Alignments and number formats
                if c_idx in [1, 2, 6, 7]:
                    cell.alignment = align_left if c_idx == 7 else align_center
                elif c_idx in [3, 4]:
                    cell.alignment = align_right
                    cell.number_format = "#,##0"
                elif c_idx == 5:
                    cell.alignment = align_right
                    cell.number_format = "0.0%"
                    
            if r_idx % 2 == 1:
                for c in range(1, 8):
                    ws_summary.cell(row=r_idx, column=c).fill = fill_zebra
            ws_summary.row_dimensions[r_idx].height = 22
            
        # Section 3: 화질별 신뢰도 분석 (A15:E19)
        ws_summary.cell(row=15, column=1, value="📐 이미지 화질 상태별 OCR 평가").font = font_section
        ws_summary.merge_cells("A15:E15")
        
        sub_headers = ["이미지 화질", "총 이미지 수", "평균 OCR 신뢰도 (Confidence)", "품질 판정"]
        for col_idx, h in enumerate(sub_headers, start=1):
            cell = ws_summary.cell(row=16, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_box
        ws_summary.row_dimensions[16].height = 25
        
        quality_rows = [
            ["저해상도 / 노이즈 있음", len(df_low_res), low_res_avg_conf, "추출 주의 (데이터 정제 필수)"],
            ["고해상도 / 노이즈 없음", len(df_normal_res), normal_res_avg_conf, "품질 우수 (원시 데이터 사용 가능)"]
        ]
        for r_idx, row_data in enumerate(quality_rows, start=17):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_data
                cell.border = border_box
                if c_idx in [1, 4]:
                    cell.alignment = align_left if c_idx == 4 else align_center
                elif c_idx == 2:
                    cell.alignment = align_right
                    cell.number_format = "#,##0"
                elif c_idx == 3:
                    cell.alignment = align_right
                    cell.number_format = "0.0%"
            ws_summary.row_dimensions[r_idx].height = 22
            
        # Column Widths Auto-fit
        for col in ws_summary.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws_summary.column_dimensions['A'].width = 18
        ws_summary.column_dimensions['G'].width = 28
        
        # Sheet 2: 영수증 금액 검증 상세 (Receipts Details)
        ws_rec = wb.create_sheet(title="영수증 금액 검증 상세")
        ws_rec.views.sheetView[0].showGridLines = True
        
        ws_rec["A1"] = "🧾 영수증 금액 추출 상세 비교 검증"
        ws_rec["A1"].font = font_section
        ws_rec.row_dimensions[1].height = 30
        
        rec_headers = ["Record ID", "카테고리", "정답 금액(GT)", "OCR 추출 금액", "일치 여부", "OCR 신뢰도", "파일 이름"]
        for c_idx, h in enumerate(rec_headers, start=1):
            cell = ws_rec.cell(row=2, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_box
        ws_rec.row_dimensions[2].height = 25
        
        for r_idx, row in df_receipts.reset_index(drop=True).iterrows():
            row_num = r_idx + 3
            is_match = (row['extracted_amount'] == row['total_amount'])
            
            ws_rec.cell(row=row_num, column=1, value=row['record_id']).alignment = align_center
            ws_rec.cell(row=row_num, column=2, value=row['category']).alignment = align_center
            
            c3 = ws_rec.cell(row=row_num, column=3, value=row['total_amount'])
            c3.alignment = align_right
            c3.number_format = "#,##0"
            
            c4 = ws_rec.cell(row=row_num, column=4, value=row['extracted_amount'] if not pd.isna(row['extracted_amount']) else "미추출")
            c4.alignment = align_right
            if isinstance(c4.value, (int, float)):
                c4.number_format = "#,##0"
            else:
                c4.font = Font(name="Malgun Gothic", size=10, bold=True, color="FF0000")
                
            c5 = ws_rec.cell(row=row_num, column=5, value="일치" if is_match else "불일치")
            c5.alignment = align_center
            c5.fill = fill_match if is_match else fill_mismatch
            
            c6 = ws_rec.cell(row=row_num, column=6, value=row['confidence'])
            c6.alignment = align_right
            c6.number_format = "0.0%"
            
            ws_rec.cell(row=row_num, column=7, value=os.path.basename(row['image_filename_ocr'])).alignment = align_left
            
            for c in range(1, 8):
                ws_rec.cell(row=row_num, column=c).font = font_data
                ws_rec.cell(row=row_num, column=c).border = border_box
            ws_rec.row_dimensions[row_num].height = 20
            
        # Sheet 3: 설문지 만족도 상세 (Surveys Details)
        ws_sur = wb.create_sheet(title="설문 만족도 점수 상세")
        ws_sur.views.sheetView[0].showGridLines = True
        
        ws_sur["A1"] = "📋 설문 부서별 및 항목 만족도 추출 상세 정합성"
        ws_sur["A1"].font = font_section
        ws_sur.row_dimensions[1].height = 30
        
        sur_headers = [
            "Record ID", "부서", 
            "만족도(GT)", "만족도(OCR)", "만족도 판정",
            "사용성(GT)", "사용성(OCR)", "사용성 판정",
            "처리속도(GT)", "처리속도(OCR)", "속도 판정"
        ]
        for c_idx, h in enumerate(sur_headers, start=1):
            cell = ws_sur.cell(row=2, column=c_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_box
        ws_sur.row_dimensions[2].height = 25
        
        for r_idx, row in enumerate(survey_details_rows):
            row_num = r_idx + 3
            
            ws_sur.cell(row=row_num, column=1, value=row['record_id']).alignment = align_center
            ws_sur.cell(row=row_num, column=2, value=row['department']).alignment = align_center
            
            # Satisfaction
            ws_sur.cell(row=row_num, column=3, value=row['satisfaction_gt']).alignment = align_right
            ws_sur.cell(row=row_num, column=4, value=row['satisfaction_ocr'] if row['satisfaction_ocr'] is not None else "누락").alignment = align_right
            c5 = ws_sur.cell(row=row_num, column=5, value=row['satisfaction_match'])
            c5.alignment = align_center
            c5.fill = fill_match if row['satisfaction_match'] == "일치" else fill_mismatch
            
            # Usability
            ws_sur.cell(row=row_num, column=6, value=row['usability_gt']).alignment = align_right
            ws_sur.cell(row=row_num, column=7, value=row['usability_ocr'] if row['usability_ocr'] is not None else "누락").alignment = align_right
            c8 = ws_sur.cell(row=row_num, column=8, value=row['usability_match'])
            c8.alignment = align_center
            c8.fill = fill_match if row['usability_match'] == "일치" else fill_mismatch
            
            # Speed
            ws_sur.cell(row=row_num, column=9, value=row['speed_gt']).alignment = align_right
            ws_sur.cell(row=row_num, column=10, value=row['speed_ocr'] if row['speed_ocr'] is not None else "누락").alignment = align_right
            c11 = ws_sur.cell(row=row_num, column=11, value=row['speed_match'])
            c11.alignment = align_center
            c11.fill = fill_match if row['speed_match'] == "일치" else fill_mismatch
            
            for c in range(1, 12):
                ws_sur.cell(row=row_num, column=c).font = font_data
                ws_sur.cell(row=row_num, column=c).border = border_box
                # Make "누락" stand out in red
                if ws_sur.cell(row=row_num, column=c).value == "누락":
                    ws_sur.cell(row=row_num, column=c).font = Font(name="Malgun Gothic", size=10, bold=True, color="FF0000")
            ws_sur.row_dimensions[row_num].height = 20
            
        # Fit column widths for other sheets as well
        for ws in [ws_rec, ws_sur]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        # Save workbook
        wb.save(REPORT_XLSX_PATH)
        print(f"💡 정합성 상세 분석 엑셀 파일이 '{REPORT_XLSX_PATH}'로 저장되었습니다.")
        
    except Exception as e:
        print(f"[경고] 엑셀 보고서 꾸미기 중 오류가 발생하여 원시 엑셀로 백업 저장합니다: {e}")
        # Fallback raw saving
        try:
            df_merged.to_excel(REPORT_XLSX_PATH, index=False)
        except Exception as e2:
            print(f"[오류] 엑셀 저장 실패: {e2}")
            
    print("======================================================")

if __name__ == "__main__":
    main()
